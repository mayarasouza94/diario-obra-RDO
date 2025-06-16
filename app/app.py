import io, os, unicodedata, traceback
from datetime import datetime, date, time
from flask import Flask, request, redirect, url_for, send_file, render_template, session, jsonify, Response
from flask.sessions import SecureCookieSessionInterface
from openpyxl import load_workbook, Workbook
import logging, sys 
import gzip

logging.basicConfig(
    level=logging.INFO,
    stream=sys.stdout,
    format="%(asctime)s %(levelname)s %(name)s: %(message)s",
)
log = logging.getLogger(__name__)


RENDERIZANDO = False

app = Flask(__name__, static_folder='static', template_folder='templates')
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024
app.secret_key = "minha_chave_secreta"

# força o Flask a usar só a sessão baseada em cookie
app.session_interface = SecureCookieSessionInterface()

ATIVIDADES_SEMANA = []
PREENCHIMENTOS_DIA = {}
ULTIMO_ENVIO = None
DATA_ULTIMO_REGISTRO = None

def load_xlsx_any(path):
    with open(path, "rb") as f:
        blob = f.read()
    if blob[:2] == b"\x1f\x8b":      # gzip?
        blob = gzip.decompress(blob)
    return load_workbook(
        filename=io.BytesIO(blob),
        data_only=True,
        read_only=True,
        keep_links=False,
    )

def remover_acentos(txt):
    return ''.join(c for c in unicodedata.normalize('NFKD', txt) if not unicodedata.combining(c))

def calcular_progresso_global():
    total_atividades = len(ATIVIDADES_SEMANA)
    preenchidas = set()
    for preenchimentos in PREENCHIMENTOS_DIA.values():
        for linha in preenchimentos:
            preenchidas.add(linha["EDT"])
    return len(preenchidas), total_atividades



@app.route('/')
def index():
    return render_template('index.html')

@app.route('/login', methods=['POST'])
def login():
    email = request.form.get('email', '').lower()
    if email.endswith('@seudominio.com.br'):
        primeiro_nome = email.split('@')[0].split('.')[0]
        session['usuario_autenticado'] = True
        session['encarregado'] = primeiro_nome.capitalize()
        session['email'] = email 
        return redirect('/diario')
    return "Acesso negado", 403

@app.route('/diario')
def diario():
    if not session.get('usuario_autenticado'):
        return redirect('/')

    global RENDERIZANDO, ATIVIDADES_SEMANA
    RENDERIZANDO = True

    xlsx_path = os.path.join(app.root_path, "prog_semanal_atual.xlsx")
    hoje = datetime.now().date()
    log.info("Planilha encontrada? %s – %s", xlsx_path, os.path.exists(xlsx_path))

    try:
        # ---------- 1) abre de forma ‘à prova de App Engine’ ------------
        with open(xlsx_path, "rb") as fh:
            blob = fh.read()
        if blob[:2] == b"\x1f\x8b":           # caso esteja compactado
            blob = gzip.decompress(blob)

        wb = load_workbook(
            filename=io.BytesIO(blob),
            data_only=True,
            read_only=True,
            keep_links=False,
        )
        sh = wb.active

        # ---------- 2) prepara cabeçalhos / linhas ----------------------
        rows_iter = sh.iter_rows(values_only=True)
        rows = list(rows_iter)
        if len(rows) < 5:
            return "Planilha com menos de 5 linhas!", 400

        headers = [str(h or "").strip() for h in rows[3]]   # linha 4 zero-based
        dados = []

        for row_values in rows[4:]:  # a partir da 5.ª linha (índice 4)
            if not any(row_values):
                continue             # pula linhas 100 % vazias
            dados.append({k: v for k, v in zip(headers, row_values)})

        # ---------- 3) utilitários de data ------------------------------
        def fmt(d):
            if isinstance(d, datetime):
                return d.strftime("%d/%m/%Y")
            if isinstance(d, str) and "-" in d:
                y, m, d_ = d.split(" ")[0].split("-")
                return f"{d_}/{m}/{y}"
            return str(d or "")

        def to_date(v):
            if isinstance(v, datetime):
                return v.date()
            try:
                return datetime.strptime(str(v).strip(), "%d/%m/%Y").date()
            except Exception:
                return None

        # ---------- 4) monta ATIVIDADES_SEMANA --------------------------
        ATIVIDADES_SEMANA.clear()
        for i, atv in enumerate(dados):
            ATIVIDADES_SEMANA.append({
                "index": i,
                "ID_Tarefa": atv.get("ID da tarefa", ""),
                "Disciplina": atv.get("Disciplina") or "Disciplina-Indefinida",
                "EDT": atv.get("EDT Cronograma", ""),
                "Descricao": atv.get("Descrição de Serviço", ""),
                "DataInicioStr": fmt(atv.get("Início Previsto")),
                "DataFimStr":   fmt(atv.get("Término Previsto")),
                "DataInicio":   to_date(atv.get("Início Previsto")),
                "DataFim":      to_date(atv.get("Término Previsto")),
                "Unidade":      atv.get("Unid Prevista  (M², M³, UNI)", ""),
                "Ferramentas":  atv.get("Ferramentas e Recursos", ""),
                "Encarregado":  atv.get("Encarregado", ""),
                "data_inicio_iso": "",
            })

        # linhas “extras”
        for extra in ["SMS", "Qualidade", "Gerais"]:
            ATIVIDADES_SEMANA.append({
                "index": len(ATIVIDADES_SEMANA),
                "ID_Tarefa": "", "Disciplina": extra,
                "EDT": extra, "Descricao": extra,
                "DataInicioStr": "", "DataFimStr": "",
                "DataInicio": hoje, "DataFim": hoje,
                "Unidade": "", "Ferramentas": "",
                "data_inicio_iso": hoje.strftime("%Y-%m-%d"),
            })

        # ---------- 5) agrupa por disciplina e calcula progresso --------
        grupos = {}
        for atv in ATIVIDADES_SEMANA:
            grupos.setdefault(atv["Disciplina"], []).append(atv)

        preenchidas, total = calcular_progresso_global()
        progresso = int(preenchidas / total * 100) if total else 0
        RENDERIZANDO = False

        encarregados = sorted(
            {
                (a.get("Encarregado") or "").strip()      # retira espaços, evita None
                for a in ATIVIDADES_SEMANA
                if a.get("Encarregado")                   # só quem realmente tem valor
            }
        )

        return render_template(
            "diario.html",
            grupos=grupos.items(),
            encarregados=encarregados,
            preenchidas=preenchidas, total=total,
            progresso=progresso,
            data_hoje=hoje.strftime("%d/%m/%Y")
        )

    except Exception:
        RENDERIZANDO = False
        log.error("ERRO_DIARIO:\n%s", traceback.format_exc())
        return "Falha interna ao ler a planilha. Verifique o arquivo e tente novamente.", 500


@app.route('/gerar', methods=['POST'])  
def gerar():
    for chave, valor in request.form.items():
        print(f"{chave}: {valor}")
    atividades   = ATIVIDADES_SEMANA
    campos       = request.form.to_dict(flat=False)
    linhas_final = []

    # lê o filtro de data do cabeçalho (YYYY-MM-DD) ou usa hoje
    raw = campos.get('fill_date', [''])[0]
    try:
        dt = datetime.strptime(raw, '%Y-%m-%d').strftime('%d/%m/%Y') if raw else datetime.now().strftime('%d/%m/%Y')
    except ValueError:
        return "Data de preenchimento inválida.", 400

    for idx_str in campos.get("atividade_idx[]", []):
            try:
                idx = int(idx_str)
                base = ATIVIDADES_SEMANA[idx]

                def get_val(chave):
                    lista = campos.get(f"{chave}[]", [])
                    return next((v for i, v in enumerate(lista)
                                if campos.get("atividade_idx[]", [])[i] == idx_str), "")

                st  = get_val("status")
                pg  = get_val("progresso")
                obs = get_val("observacoes")

                linha = {
                    "EDT":                       base.get("EDT",""),
                    "TAREFA":                    base.get("Descricao",""),
                    "STATUS":                    st,
                    "DATA DE INICIO (PROG)":     base.get("DataInicioStr",""),
                    "DATA DE INICIO REAL":       dt if st=="Iniciada" else "",
                    "DATA DE TERMINO (PROG)":    base.get("DataFimStr",""),
                    "DATA DE TERMINO REAL":      dt if st=="Finalizada" else "",
                    "PROGRESSO":                 pg,
                    "Unidade de medida":         base.get("Unidade",""),
                    "ENCARREGADO":               base.get("Encarregado",""),
                    "Disciplina":                base.get("Disciplina",""),
                    "OBS":                       obs,
                    "RECURSOS":                  ", ".join(campos.get(f"ferramentas_recursos_{idx}[]", [])),
                    "ATIVIDADES PRELIMINARES":   campos.get(f"subtarefas_{idx}_nome[]", []),
                    "DATA DE PREENCHIMENTO":     dt.strftime('%d/%m/%Y'),
                    "USUÁRIO":                   session.get("email","desconhecido@seudominio.com.br")
                }

                linhas_final.append(linha)

            except Exception as e:
                log.warning(f"Erro ao processar atividade {idx_str}: {e}")
                continue

    if not linhas_final:
        return "Nenhuma atividade válida foi recebida", 400

    user       = session.get('encarregado','desconhecido')
    if user not in PREENCHIMENTOS_DIA:
        PREENCHIMENTOS_DIA[user] = []

    antigos = {l["EDT"] for l in linhas_final}
    PREENCHIMENTOS_DIA[user] = [
        l for l in PREENCHIMENTOS_DIA[user] if l["EDT"] not in antigos
    ] + linhas_final

    log.info("Preenchimentos salvos: %d", len(PREENCHIMENTOS_DIA[user]))
    return redirect(url_for('diario'))

# ——— exporta o Excel só dos registros de HOJE ———

@app.route('/exportar_diario')
def exportar_diario():
    if not session.get('usuario_autenticado'):
        return redirect(url_for('index'))

    usuario = session['encarregado']

    print("======== PREENCHIMENTOS DO USUÁRIO =========")
    for r in PREENCHIMENTOS_DIA.get(usuario, []):
        print(r)

    # lê ?date=YYYY-MM-DD da query string; se não vier, usa hoje
    date_str = request.args.get('date')
    try:
        dt = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else date.today()
    except ValueError:
        return "Formato de data inválido. Use YYYY-MM-DD.", 400

    filtro = dt.strftime('%d/%m/%Y')
    registros = []

    for r in PREENCHIMENTOS_DIA.get(usuario, []):
        data_preench = r.get("DATA DE PREENCHIMENTO")
        data_obj = None
        # tenta os dois formatos
        for fmt in ('%d/%m/%Y', '%Y-%m-%d'):
            try:
                data_obj = datetime.strptime(data_preench, fmt).date()
                break
            except Exception:
                continue
        if data_obj == dt:
            registros.append(r)

    if not registros:
        return f"Nenhum registro para {filtro}.", 400

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Diário")
    ws.append([
        "EDT", "Tarefa", "Status", "Data", "Progresso", "Atividades Preliminares", "Observações"
    ])
    for r in registros:
        prelim = "; ".join(r.get("ATIVIDADES PRELIMINARES", []))
        ws.append([
            r.get("EDT", ""),
            r.get("TAREFA", ""),
            r.get("STATUS", ""),
            r.get("DATA DE PREENCHIMENTO", ""),
            r.get("PROGRESSO", ""),
            "; ".join(r.get("ATIVIDADES PRELIMINARES", [])),
            r.get("OBS", ""),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"diario_{dt.strftime('%Y-%m-%d')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
@app.route('/carregando')
def carregando():
    return "ok" if not RENDERIZANDO else "carregando"

@app.route('/salvar_tarefa', methods=['POST'])
def salvar_tarefa():
    if not session.get('usuario_autenticado'):
        return "Não autenticado", 403

    try:
        dados = request.get_json()
        usuario = session.get("encarregado", "desconhecido")

        if usuario not in PREENCHIMENTOS_DIA:
            PREENCHIMENTOS_DIA[usuario] = []

        dados["atividade_idx"] = dados.get("atividade_idx") or str(len(PREENCHIMENTOS_DIA[usuario]))

        # Remove tarefas com mesma atividade_idx e data
        idx_atual = dados.get("atividade_idx")
        data_atual = dados.get("DATA DE PREENCHIMENTO")
        PREENCHIMENTOS_DIA[usuario] = [
            t for t in PREENCHIMENTOS_DIA[usuario]
            if not (t.get("atividade_idx") == idx_atual and t.get("DATA DE PREENCHIMENTO") == data_atual)
        ]

        # Garante que "ATIVIDADES PRELIMINARES" seja lista
        preliminares = dados.get("ATIVIDADES PRELIMINARES") or []
        if isinstance(preliminares, str):
            preliminares = [preliminares]
        elif isinstance(preliminares, list):
            preliminares = [str(p).strip() for p in preliminares if str(p).strip()]
        else:
            preliminares = []

        dados["ATIVIDADES PRELIMINARES"] = preliminares

        PREENCHIMENTOS_DIA[usuario].append(dados)
        print(f"[salvar_tarefa] Tarefa {dados.get('EDT')} salva para {usuario}")
        return jsonify({"ok": True})

    except Exception as e:
        print("Erro ao salvar tarefa:", e)
        return jsonify({"erro": str(e)}), 500
        
if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=8080)

